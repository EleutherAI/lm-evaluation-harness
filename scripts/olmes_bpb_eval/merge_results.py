"""Merge Ray shard JSONs into reviewable CSV and XLSX result tables."""

from __future__ import annotations

import argparse
import csv
import json
import math
from collections import defaultdict
from pathlib import Path
from statistics import fmean
from typing import Any

from lm_eval.tasks import TaskManager


try:
    from scripts.olmes_bpb_eval.ray_driver import resolve_leaves
except ModuleNotFoundError:  # direct ``python scripts/.../merge_results.py``
    from ray_driver import resolve_leaves


BPB_METRICS = {
    "bpb_macro",
    "bits_per_byte_corr",
    "bpb_corpus",
    "bpb_total_loglikelihood",
    "bpb_total_bytes",
}
BOOKKEEPING_RESULT_KEYS = {"alias", "name", "sample_count", "sample_len"}
PER_VARIANT_NOTE = "See Task Metrics for source GPQA variants"
CONSTRAINTS = Path(__file__).with_name("constraints.txt")


def pinned_runtime_versions(path: Path = CONSTRAINTS) -> dict[str, str]:
    """Read the exact package pins enforced by the campaign installer."""
    versions = {}
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.split("#", 1)[0].strip()
        if not line:
            continue
        if "==" not in line:
            raise ValueError(f"Campaign constraint is not an exact pin: {line!r}")
        package, version = line.split("==", 1)
        versions[package] = version
    return versions


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--results-dir", type=Path, required=True)
    parser.add_argument(
        "--suite",
        type=Path,
        default=Path(__file__).with_name("suite.json"),
    )
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument(
        "--allow-incomplete",
        action="store_true",
        help=(
            "Export an explicitly partial run, such as a code-blocked phase. "
            "Unexpected models/tasks, wrong revisions, and truncated tasks are "
            "still rejected."
        ),
    )
    return parser.parse_args()


def model_name(result: dict[str, Any]) -> str:
    args = result.get("config", {}).get("model_args")
    if isinstance(args, dict) and args.get("pretrained"):
        return args["pretrained"]
    return result.get("config", {}).get("model", "unknown")


def metric_parts(key: str) -> tuple[str, str]:
    return tuple(key.split(",", 1)) if "," in key else (key, "none")


def task_family_map(suite: dict[str, Any]) -> dict[str, str]:
    manager = TaskManager()
    mapping = {}
    for family in suite["families"]:
        for task in resolve_leaves(manager, family["selector"]):
            previous = mapping.setdefault(task, family["name"])
            if previous != family["name"]:
                raise ValueError(f"Task {task} belongs to two suite families")
    return mapping


def load_results(results_dir: Path) -> list[dict[str, Any]]:
    loaded = []
    for path in sorted(results_dir.glob("Qwen__*/**/*.json")):
        result = json.loads(path.read_text(encoding="utf-8"))
        if "results" not in result:
            continue
        result["_path"] = str(path)
        loaded.append(result)
    if not loaded:
        raise FileNotFoundError(f"No shard result JSONs found under {results_dir}")
    return loaded


def _dataset_revision_status(
    result: dict[str, Any],
    task: str,
    suite: dict[str, Any],
) -> str:
    config = result.get("configs", {}).get(task)
    if not isinstance(config, dict):
        return "UNVERIFIED: result JSON has no serialized task config"

    dataset_path = config.get("dataset_path")
    dataset_kwargs = config.get("dataset_kwargs")
    actual_revision = (
        dataset_kwargs.get("revision") if isinstance(dataset_kwargs, dict) else None
    )
    expected_revision = suite.get("dataset_revisions", {}).get(dataset_path)
    if expected_revision is not None:
        if actual_revision is None:
            return (
                f"UNVERIFIED: {dataset_path} config does not expose a revision; "
                f"suite expects {expected_revision}"
            )
        if actual_revision != expected_revision:
            raise ValueError(
                f"{task}: dataset {dataset_path} revision {actual_revision!r} does "
                f"not match suite pin {expected_revision!r}"
            )
        return f"Verified dataset revision: {actual_revision}"

    metadata = config.get("metadata")
    dataset_repository = (
        metadata.get("dataset_repository") if isinstance(metadata, dict) else None
    )
    metadata_revision = (
        metadata.get("dataset_revision") if isinstance(metadata, dict) else None
    )
    if dataset_path == "json" and dataset_repository in suite.get(
        "dataset_revisions", {}
    ):
        expected_revision = suite["dataset_revisions"][dataset_repository]
        if metadata_revision != expected_revision:
            raise ValueError(
                f"{task}: raw dataset metadata revision {metadata_revision!r} does "
                f"not match suite pin {expected_revision!r}"
            )
        data_files = dataset_kwargs.get("data_files") if dataset_kwargs else None
        serialized_data_files = json.dumps(data_files, sort_keys=True)
        if f"/resolve/{expected_revision}/" not in serialized_data_files:
            raise ValueError(
                f"{task}: raw data URL does not embed pinned revision "
                f"{expected_revision!r}"
            )
        return f"Verified dataset revision: {expected_revision}"

    # Local JSON tasks such as BBEH expose their pinned upstream source revision
    # in metadata instead of a Hugging Face dataset revision.
    source_revision = (
        metadata.get("source_revision") if isinstance(metadata, dict) else None
    )
    source_pins = set(suite.get("source_revisions", {}).values())
    if dataset_path == "json" and source_revision in source_pins:
        return f"Verified source revision: {source_revision}"

    if actual_revision is not None:
        raise ValueError(
            f"{task}: result exposes dataset {dataset_path!r} revision "
            f"{actual_revision!r}, but suite.json has no pin for that dataset"
        )
    return (
        f"UNVERIFIED: no comparable suite dataset/source pin for "
        f"dataset_path={dataset_path!r}"
    )


def validate_results(
    results: list[dict[str, Any]],
    suite: dict[str, Any],
    family_by_task: dict[str, str],
    *,
    allow_incomplete: bool = False,
) -> None:
    """Validate exact release identity and completeness before exporting.

    ``allow_incomplete`` permits missing expected models/tasks only. It never
    permits unexpected rows, revision mismatches, duplicate task shards, or a
    present task evaluated on fewer than its full source samples.
    """

    expected_models = {model["name"]: model["revision"] for model in suite["models"]}
    expected_tasks = set(family_by_task)
    observed_by_model: defaultdict[str, set[str]] = defaultdict(set)
    expected_runtime_versions = pinned_runtime_versions()
    source_identities: set[tuple[str, str]] = set()

    for result in results:
        model = model_name(result)
        if model not in expected_models:
            raise ValueError(f"Unexpected result model: {model!r}")

        model_args = result.get("config", {}).get("model_args")
        if not isinstance(model_args, dict):
            raise TypeError(f"{model}: result does not expose model_args")
        expected_revision = expected_models[model]
        for key in ("revision", "tokenizer_revision"):
            actual_revision = model_args.get(key)
            if actual_revision != expected_revision:
                raise ValueError(
                    f"{model}: {key} {actual_revision!r} does not match suite pin "
                    f"{expected_revision!r}"
                )

        runtime = result.get("olmes_bpb_runtime")
        if not isinstance(runtime, dict):
            raise TypeError(f"{model}: result does not expose an OLMES BPB runtime")
        if runtime.get("git_dirty") is not False:
            raise ValueError(f"{model}: result was produced from a dirty Git tree")
        git_commit = runtime.get("git_commit")
        source_snapshot = runtime.get("source_snapshot_sha256")
        if not isinstance(git_commit, str) or not git_commit:
            raise ValueError(f"{model}: result has no lm-eval Git commit")
        if not isinstance(source_snapshot, str) or not source_snapshot:
            raise ValueError(f"{model}: result has no lm-eval source snapshot")
        source_identities.add((git_commit, source_snapshot))
        actual_versions = runtime.get("packages")
        if not isinstance(actual_versions, dict):
            raise TypeError(f"{model}: result does not expose runtime package versions")
        mismatches = {
            package: (actual_versions.get(package), expected)
            for package, expected in expected_runtime_versions.items()
            if actual_versions.get(package) != expected
        }
        if mismatches:
            raise ValueError(
                f"{model}: runtime package versions do not match constraints: "
                f"{mismatches}"
            )

        result_tasks = set(result.get("results", {}))
        unexpected_tasks = result_tasks - expected_tasks
        if unexpected_tasks:
            raise ValueError(
                f"{model}: unexpected result task(s): {sorted(unexpected_tasks)}"
            )
        duplicated_tasks = result_tasks.intersection(observed_by_model[model])
        if duplicated_tasks:
            raise ValueError(
                f"{model}: duplicate result task(s) across shards: "
                f"{sorted(duplicated_tasks)}"
            )

        sample_counts = result.get("n-samples", {})
        revision_statuses = result.setdefault("_dataset_revision_status", {})
        for task in result_tasks:
            counts = sample_counts.get(task)
            if not isinstance(counts, dict):
                raise TypeError(
                    f"{model}/{task}: cannot prove completeness because n-samples "
                    "does not expose effective and original counts"
                )
            effective = counts.get("effective")
            original = counts.get("original")
            if effective is None or original is None:
                raise ValueError(
                    f"{model}/{task}: n-samples must expose effective and original"
                )
            if effective != original:
                raise ValueError(
                    f"{model}/{task}: truncated result has {effective}/{original} "
                    "effective samples"
                )
            revision_statuses[task] = _dataset_revision_status(result, task, suite)
        observed_by_model[model].update(result_tasks)

    if len(source_identities) != 1:
        raise ValueError(
            "Result shards were produced from multiple lm-eval source identities: "
            f"{sorted(source_identities)}"
        )

    if allow_incomplete:
        return

    missing_models = set(expected_models) - set(observed_by_model)
    if missing_models:
        raise ValueError(f"Missing expected result model(s): {sorted(missing_models)}")
    missing_tasks = {
        model: sorted(expected_tasks - observed_by_model[model])
        for model in expected_models
        if expected_tasks - observed_by_model[model]
    }
    if missing_tasks:
        details = "; ".join(
            f"{model}: {len(tasks)} missing ({', '.join(tasks[:5])}"
            f"{'...' if len(tasks) > 5 else ''})"
            for model, tasks in missing_tasks.items()
        )
        raise ValueError(f"Missing expected task results: {details}")


def extract_rows(
    results: list[dict[str, Any]], family_by_task: dict[str, str]
) -> list[dict[str, Any]]:
    rows = []
    seen: set[tuple[str, str, str, str]] = set()
    for result in results:
        model = model_name(result)
        sample_counts = result.get("n-samples", {})
        for task, metrics in result["results"].items():
            family = family_by_task.get(task)
            if family is None:
                continue
            count = sample_counts.get(task, {})
            if isinstance(count, dict):
                count = count.get("effective", count.get("original"))
            for key, value in metrics.items():
                metric, filter_name = metric_parts(key)
                if metric in BOOKKEEPING_RESULT_KEYS or metric.endswith("_stderr"):
                    continue
                if not isinstance(value, (int, float)) or isinstance(value, bool):
                    continue
                identity = (model, task, metric, filter_name)
                if identity in seen:
                    raise ValueError(f"Duplicate task metric across shards: {identity}")
                seen.add(identity)
                stderr_key = f"{metric}_stderr,{filter_name}"
                rows.append(
                    {
                        "Model": model,
                        "Family": family,
                        "Task": task,
                        "Metric": metric,
                        "Filter": filter_name,
                        "Value": float(value),
                        "StdErr": metrics.get(stderr_key),
                        "Samples": count,
                        "Dataset Revision Status": result.get(
                            "_dataset_revision_status", {}
                        ).get(
                            task,
                            "UNVERIFIED: validate_results was not run",
                        ),
                        "Result File": result["_path"],
                    }
                )
    return rows


def _weighted(rows: list[dict[str, Any]]) -> float | None:
    usable = [row for row in rows if isinstance(row["Samples"], (int, float))]
    if not usable:
        return fmean(row["Value"] for row in rows) if rows else None
    total = sum(row["Samples"] for row in usable)
    return (
        sum(row["Value"] * row["Samples"] for row in usable) / total if total else None
    )


def _bbeh_adjusted_harmonic(rows: list[dict[str, Any]]) -> float | None:
    if not rows:
        return None
    return len(rows) / math.fsum(1.0 / (row["Value"] + 0.01) for row in rows)


def _one_metric_row_per_task(
    rows: list[dict[str, Any]], metric: str
) -> list[dict[str, Any]]:
    by_task: defaultdict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        if row["Metric"] == metric:
            by_task[row["Task"]].append(row)
    selected = []
    for task_rows in by_task.values():
        selected.append(
            min(task_rows, key=lambda row: (row["Filter"] != "none", row["Filter"]))
        )
    return selected


def _original_rows(
    family_rows: list[dict[str, Any]], family: dict[str, Any]
) -> tuple[str | None, list[dict[str, Any]]]:
    metric_by_suffix = family.get("primary_metric_by_task_suffix")
    if metric_by_suffix:
        selected = []
        labels = []
        for suffix, metric in metric_by_suffix.items():
            matching = [row for row in family_rows if row["Task"].endswith(suffix)]
            selected.extend(_one_metric_row_per_task(matching, metric))
            labels.append(f"{metric} ({suffix})")
        return "; ".join(labels), selected

    for candidate in family["primary_metrics"]:
        candidate_rows = _one_metric_row_per_task(family_rows, candidate)
        if candidate_rows:
            return candidate, candidate_rows
    return None, []


def build_summary(
    rows: list[dict[str, Any]], suite: dict[str, Any]
) -> list[dict[str, Any]]:
    manager = TaskManager()
    expected_by_family = {
        family["name"]: set(resolve_leaves(manager, family["selector"]))
        for family in suite["families"]
    }
    grouped: defaultdict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        grouped[(row["Model"], row["Family"])].append(row)

    summary = []
    for model in [model["name"] for model in suite["models"]]:
        for family in suite["families"]:
            family_rows = grouped.get((model, family["name"]), [])
            original_metric, original_rows = _original_rows(family_rows, family)

            aggregation = family["aggregation"]
            if aggregation == "single":
                original_value = original_rows[0]["Value"] if original_rows else None
            elif aggregation == "macro_task":
                original_value = (
                    fmean(row["Value"] for row in original_rows)
                    if original_rows
                    else None
                )
            elif aggregation == "bbeh_adjusted_harmonic":
                original_value = _bbeh_adjusted_harmonic(original_rows)
            elif aggregation == "per_variant":
                original_value = None
            else:
                original_value = _weighted(original_rows)

            macro_rows = _one_metric_row_per_task(family_rows, "bpb_macro")
            total_ll_rows = _one_metric_row_per_task(
                family_rows, "bpb_total_loglikelihood"
            )
            total_byte_rows = _one_metric_row_per_task(family_rows, "bpb_total_bytes")
            bpb_macro = _weighted(macro_rows)
            total_ll = sum(row["Value"] for row in total_ll_rows)
            total_bytes = sum(row["Value"] for row in total_byte_rows)
            bpb_corpus = (
                -total_ll / (total_bytes * math.log(2)) if total_bytes else None
            )
            expected_tasks = expected_by_family[family["name"]]
            observed_tasks = {row["Task"] for row in family_rows}
            original_tasks = {row["Task"] for row in original_rows}
            bpb_tasks = {row["Task"] for row in macro_rows}
            bpb_total_tasks = {row["Task"] for row in total_ll_rows}.intersection(
                row["Task"] for row in total_byte_rows
            )
            missing_parts = []
            if expected_tasks - observed_tasks:
                missing_parts.append(
                    f"task rows {len(observed_tasks)}/{len(expected_tasks)}"
                )
            if expected_tasks - original_tasks:
                missing_parts.append(
                    f"original metrics {len(original_tasks)}/{len(expected_tasks)}"
                )
            if family.get("compute_bpb", True):
                if expected_tasks - bpb_tasks:
                    missing_parts.append(
                        f"macro BPB {len(bpb_tasks)}/{len(expected_tasks)}"
                    )
                if expected_tasks - bpb_total_tasks:
                    missing_parts.append(
                        f"corpus BPB totals {len(bpb_total_tasks)}/{len(expected_tasks)}"
                    )
            revision_rows = [
                row for row in family_rows if "Dataset Revision Status" in row
            ]
            unverified_revision_tasks = {
                row["Task"]
                for row in revision_rows
                if not row["Dataset Revision Status"].startswith("Verified ")
            }
            if unverified_revision_tasks:
                missing_parts.append(
                    "dataset provenance unverified for "
                    f"{len(unverified_revision_tasks)}/{len(expected_tasks)} tasks"
                )
            status = (
                "Complete"
                if not missing_parts
                else "Incomplete: " + ", ".join(missing_parts)
            )
            if not family.get("compute_bpb", True):
                bpb_macro = None
                bpb_corpus = None
                bpb_status = family.get("bpb_status", "BPB N/A")
                status = bpb_status if not missing_parts else f"{bpb_status}; {status}"
            if aggregation == "per_variant" and family_rows:
                status = f"{status}; {PER_VARIANT_NOTE}"

            summary.append(
                {
                    "Model": model,
                    "Family": family["name"],
                    "Selector": family["selector"],
                    "Original Metric": original_metric,
                    "Original Value": original_value,
                    "Original Aggregation": aggregation,
                    "BPB Macro": bpb_macro,
                    "BPB Corpus": bpb_corpus,
                    "Leaf Tasks": len({row["Task"] for row in family_rows}),
                    "Status / Notes": status,
                }
            )
    return summary


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)


def manifest_rows(
    suite: dict[str, Any], results: list[dict[str, Any]]
) -> list[dict[str, str]]:
    rows = [
        {"Category": "suite", "Name": "suite_id", "Pinned Value": suite["suite_id"]},
        {
            "Category": "suite",
            "Name": "excluded",
            "Pinned Value": ", ".join(suite["excluded"]),
        },
    ]
    for model in suite["models"]:
        rows.append(
            {
                "Category": "model",
                "Name": model["name"],
                "Pinned Value": model["revision"],
            }
        )
    for name, revision in suite["dataset_revisions"].items():
        rows.append({"Category": "dataset", "Name": name, "Pinned Value": revision})
    for name, revision in suite["source_revisions"].items():
        rows.append({"Category": "source", "Name": name, "Pinned Value": revision})
    runtimes = {
        json.dumps(result.get("olmes_bpb_runtime", {}), sort_keys=True)
        for result in results
    }
    for index, runtime in enumerate(sorted(runtimes), start=1):
        rows.append(
            {"Category": "runtime", "Name": f"runtime_{index}", "Pinned Value": runtime}
        )
    actual_model_revisions = {
        (
            model_name(result),
            result.get("config", {}).get("model_args", {}).get("revision"),
            result.get("config", {}).get("model_args", {}).get("tokenizer_revision"),
        )
        for result in results
    }
    for model, revision, tokenizer_revision in sorted(actual_model_revisions):
        rows.extend(
            [
                {
                    "Category": "result_model_revision",
                    "Name": model,
                    "Pinned Value": revision,
                },
                {
                    "Category": "result_tokenizer_revision",
                    "Name": model,
                    "Pinned Value": tokenizer_revision,
                },
            ]
        )
    for result in results:
        model = model_name(result)
        for task, status in sorted(result.get("_dataset_revision_status", {}).items()):
            if not status.startswith("Verified "):
                rows.append(
                    {
                        "Category": "dataset_provenance",
                        "Name": f"{model}:{task}",
                        "Pinned Value": status,
                    }
                )
    return rows


def validate_summary_for_release(
    summary: list[dict[str, Any]], suite: dict[str, Any]
) -> None:
    family_by_name = {family["name"]: family for family in suite["families"]}
    incomplete = []
    for row in summary:
        family = family_by_name[row["Family"]]
        expected_status = (
            family.get("bpb_status", "BPB N/A")
            if not family.get("compute_bpb", True)
            else "Complete"
        )
        if family["aggregation"] == "per_variant":
            expected_status = f"{expected_status}; {PER_VARIANT_NOTE}"
        if row["Status / Notes"] != expected_status:
            incomplete.append(
                f"{row['Model']}/{row['Family']}: {row['Status / Notes']}"
            )
    if incomplete:
        raise ValueError(
            "Release export is incomplete; rerun with --allow-incomplete only for "
            "an explicitly partial review workbook: " + "; ".join(incomplete)
        )


def write_xlsx(
    path: Path,
    summary: list[dict[str, Any]],
    detail: list[dict[str, Any]],
    manifest: list[dict[str, Any]],
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    workbook = Workbook()
    workbook.remove(workbook.active)
    sheets = (
        ("Family Summary", summary),
        ("Task Metrics", detail),
        ("Run Manifest", manifest),
    )
    header_fill = PatternFill("solid", fgColor="E2E8F0")
    header_font = Font(bold=True, color="1F2937")
    border = Border(bottom=Side(style="thin", color="CBD5E1"))
    for title, data in sheets:
        sheet = workbook.create_sheet(title)
        headers = list(data[0]) if data else []
        sheet.append(headers)
        for row in data:
            sheet.append([row.get(header) for header in headers])
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(vertical="top")
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column in sheet.columns:
            letter = column[0].column_letter
            width = min(max(len(str(cell.value or "")) for cell in column) + 2, 60)
            sheet.column_dimensions[letter].width = max(width, 12)
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = "0.000000"
                cell.alignment = Alignment(
                    vertical="top", wrap_text=title == "Run Manifest"
                )
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def main() -> None:
    args = parse_args()
    suite = json.loads(args.suite.read_text(encoding="utf-8"))
    results = load_results(args.results_dir)
    family_by_task = task_family_map(suite)
    validate_results(
        results,
        suite,
        family_by_task,
        allow_incomplete=args.allow_incomplete,
    )
    detail = extract_rows(results, family_by_task)
    summary = build_summary(detail, suite)
    if not args.allow_incomplete:
        validate_summary_for_release(summary, suite)
    manifest = manifest_rows(suite, results)
    write_csv(args.output.with_suffix(".summary.csv"), summary)
    write_csv(args.output.with_suffix(".task_metrics.csv"), detail)
    write_xlsx(args.output, summary, detail, manifest)


if __name__ == "__main__":
    main()
