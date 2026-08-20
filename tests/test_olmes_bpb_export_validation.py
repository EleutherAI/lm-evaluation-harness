import copy

import pytest

from scripts.olmes_bpb_eval.merge_results import (
    build_summary,
    extract_rows,
    manifest_rows,
    pinned_runtime_versions,
    validate_results,
    validate_summary_for_release,
    write_xlsx,
)


MODEL = "Qwen/Qwen3.5-0.8B-Base"
MODEL_REVISION = "model-revision"
DATASET_REVISION = "dataset-revision"
FAMILY_BY_TASK = {"arc_easy": "arc_easy"}


def suite():
    return {
        "suite_id": "test-suite",
        "excluded": [],
        "models": [{"name": MODEL, "revision": MODEL_REVISION}],
        "families": [
            {
                "name": "arc_easy",
                "selector": "arc_easy",
                "primary_metrics": ["acc_norm", "acc"],
                "aggregation": "single",
            }
        ],
        "dataset_revisions": {"allenai/ai2_arc": DATASET_REVISION},
        "source_revisions": {},
    }


def result(*, effective=2376, original=2376, dataset_revision=DATASET_REVISION):
    dataset_kwargs = {}
    if dataset_revision is not None:
        dataset_kwargs["revision"] = dataset_revision
    return {
        "_path": "/results/arc.json",
        "config": {
            "model": "vllm",
            "model_args": {
                "pretrained": MODEL,
                "revision": MODEL_REVISION,
                "tokenizer_revision": MODEL_REVISION,
            },
        },
        "configs": {
            "arc_easy": {
                "dataset_path": "allenai/ai2_arc",
                "dataset_kwargs": dataset_kwargs,
            }
        },
        "results": {
            "arc_easy": {
                "alias": "arc_easy",
                "acc,none": 0.5,
                "acc_norm,none": 0.6,
                "bpb_macro,none": 1.25,
                "bpb_total_loglikelihood,none": -10.0,
                "bpb_total_bytes,none": 20,
                "sample_len": effective,
            }
        },
        "n-samples": {"arc_easy": {"effective": effective, "original": original}},
        "olmes_bpb_runtime": {
            "git_commit": "commit",
            "git_dirty": False,
            "source_snapshot_sha256": "snapshot",
            "packages": pinned_runtime_versions(),
        },
    }


def test_validation_accepts_exact_pins_and_full_sample_count():
    shard = result()

    validate_results([shard], suite(), FAMILY_BY_TASK)

    assert shard["_dataset_revision_status"]["arc_easy"] == (
        f"Verified dataset revision: {DATASET_REVISION}"
    )


def test_truncated_task_is_rejected_even_for_incomplete_export():
    with pytest.raises(ValueError, match="truncated result has 2/2376"):
        validate_results(
            [result(effective=2)],
            suite(),
            FAMILY_BY_TASK,
            allow_incomplete=True,
        )


@pytest.mark.parametrize("revision_key", ["revision", "tokenizer_revision"])
def test_wrong_model_or_tokenizer_revision_is_rejected(revision_key):
    shard = result()
    shard["config"]["model_args"][revision_key] = "wrong-revision"

    with pytest.raises(ValueError, match=revision_key):
        validate_results([shard], suite(), FAMILY_BY_TASK)


def test_unexpected_model_is_rejected():
    shard = result()
    shard["config"]["model_args"]["pretrained"] = "unexpected/model"

    with pytest.raises(ValueError, match="Unexpected result model"):
        validate_results([shard], suite(), FAMILY_BY_TASK)


def test_unexpected_task_is_rejected():
    shard = result()
    shard["results"]["unexpected_task"] = {"acc,none": 1.0}

    with pytest.raises(ValueError, match="unexpected result task"):
        validate_results([shard], suite(), FAMILY_BY_TASK)


def test_missing_expected_task_requires_explicit_incomplete_export():
    expected_tasks = {**FAMILY_BY_TASK, "arc_challenge": "arc_challenge"}

    with pytest.raises(ValueError, match="Missing expected task results"):
        validate_results([result()], suite(), expected_tasks)

    validate_results([result()], suite(), expected_tasks, allow_incomplete=True)


def test_sample_len_is_not_exported_as_a_task_metric():
    shard = result()
    validate_results([shard], suite(), FAMILY_BY_TASK)

    rows = extract_rows([shard], FAMILY_BY_TASK)

    assert "sample_len" not in {row["Metric"] for row in rows}
    assert {row["Dataset Revision Status"] for row in rows} == {
        f"Verified dataset revision: {DATASET_REVISION}"
    }


def test_wrong_dataset_revision_is_rejected():
    with pytest.raises(ValueError, match="does not match suite pin"):
        validate_results(
            [result(dataset_revision="wrong-revision")],
            suite(),
            FAMILY_BY_TASK,
        )


def test_pinned_raw_json_dataset_url_is_verified():
    shard = result(dataset_revision=None)
    shard["configs"]["arc_easy"] = {
        "dataset_path": "json",
        "dataset_kwargs": {
            "data_files": {
                "validation": (
                    "https://huggingface.co/datasets/allenai/ai2_arc/resolve/"
                    f"{DATASET_REVISION}/validation.json"
                )
            }
        },
        "metadata": {
            "dataset_repository": "allenai/ai2_arc",
            "dataset_revision": DATASET_REVISION,
        },
    }

    validate_results([shard], suite(), FAMILY_BY_TASK)

    assert shard["_dataset_revision_status"]["arc_easy"] == (
        f"Verified dataset revision: {DATASET_REVISION}"
    )


def test_missing_dataset_revision_is_visible_and_not_release_complete():
    shard = result(dataset_revision=None)
    validate_results([shard], suite(), FAMILY_BY_TASK)
    rows = extract_rows([shard], FAMILY_BY_TASK)

    summary = build_summary(rows, suite())

    assert "dataset provenance unverified" in summary[0]["Status / Notes"]
    with pytest.raises(ValueError, match="Release export is incomplete"):
        validate_summary_for_release(summary, suite())
    provenance = [
        row
        for row in manifest_rows(suite(), [shard])
        if row["Category"] == "dataset_provenance"
    ]
    assert len(provenance) == 1
    assert "UNVERIFIED" in provenance[0]["Pinned Value"]


def test_duplicate_task_across_shards_is_rejected():
    second_shard = copy.deepcopy(result())
    second_shard["_path"] = "/results/arc-duplicate.json"

    with pytest.raises(ValueError, match="duplicate result task"):
        validate_results([result(), second_shard], suite(), FAMILY_BY_TASK)


def test_dirty_or_mixed_source_snapshots_are_rejected():
    dirty = result()
    dirty["olmes_bpb_runtime"]["git_dirty"] = True
    with pytest.raises(ValueError, match="dirty Git tree"):
        validate_results([dirty], suite(), FAMILY_BY_TASK)

    second_shard = copy.deepcopy(result())
    second_shard["results"] = {}
    second_shard["n-samples"] = {}
    second_shard["olmes_bpb_runtime"]["source_snapshot_sha256"] = "other"
    with pytest.raises(ValueError, match="multiple lm-eval source identities"):
        validate_results(
            [result(), second_shard],
            suite(),
            FAMILY_BY_TASK,
            allow_incomplete=True,
        )


def test_wrong_runtime_package_version_is_rejected():
    shard = result()
    shard["olmes_bpb_runtime"]["packages"]["transformers"] = "wrong"

    with pytest.raises(ValueError, match="do not match constraints"):
        validate_results([shard], suite(), FAMILY_BY_TASK)


def test_complete_per_variant_note_is_release_valid_but_does_not_mask_missing():
    per_variant_suite = suite()
    per_variant_suite["families"][0]["aggregation"] = "per_variant"
    complete = [
        {
            "Model": MODEL,
            "Family": "arc_easy",
            "Status / Notes": ("Complete; See Task Metrics for source GPQA variants"),
        }
    ]

    validate_summary_for_release(complete, per_variant_suite)

    incomplete = copy.deepcopy(complete)
    incomplete[0]["Status / Notes"] = (
        "Incomplete: original metrics 0/1; See Task Metrics for source GPQA variants"
    )
    with pytest.raises(ValueError, match="Release export is incomplete"):
        validate_summary_for_release(incomplete, per_variant_suite)


def test_workbook_has_review_tabs_and_exact_column_order(tmp_path):
    from openpyxl import load_workbook

    summary = [
        {
            "Model": MODEL,
            "Family": "arc_easy",
            "Selector": "arc_easy",
            "Original Metric": "acc_norm",
            "Original Value": 0.5,
            "Original Aggregation": "single",
            "BPB Macro": 1.0,
            "BPB Corpus": 0.9,
            "Leaf Tasks": 1,
            "Status / Notes": "Complete",
        }
    ]
    shard = result()
    validate_results([shard], suite(), FAMILY_BY_TASK)
    detail = extract_rows([shard], FAMILY_BY_TASK)
    manifest = manifest_rows(suite(), [shard])
    output = tmp_path / "results.xlsx"

    write_xlsx(output, summary, detail, manifest)

    workbook = load_workbook(output, read_only=True)
    assert workbook.sheetnames == ["Family Summary", "Task Metrics", "Run Manifest"]
    assert [cell.value for cell in workbook["Family Summary"][1]] == list(summary[0])
    assert [cell.value for cell in workbook["Task Metrics"][1]] == list(detail[0])
    assert [cell.value for cell in workbook["Run Manifest"][1]] == list(manifest[0])
